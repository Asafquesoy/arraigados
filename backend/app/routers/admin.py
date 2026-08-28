import csv
import io
import os
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import case, func, or_
from sqlalchemy.orm import Session, joinedload

from ..config import settings
from ..database import get_db
from ..models import AdminRole, AdminUser, AppSettings, Camper, Equipo, Sexo, TallaCamisa, TipoParticipante, Zona
from ..schemas import (
    AdminUserCreate,
    AdminUserOut,
    AdminUserUpdate,
    AppSettingsOut,
    AppSettingsUpdate,
    AsistenciaStatsResponse,
    AsistenciaUpdate,
    CamperListResponse,
    CamperOut,
    CamperUpdate,
    ComprobanteStatsResponse,
    EquipoAsignacion,
    PagoUpdate,
    TallaStatsItem,
    TallaStatsResponse,
)
from ..security import get_current_admin, hash_password, require_role
from ..validacion_camper import normalizar_datos_camper, validar_datos_camper

router = APIRouter(prefix="/api/admin", tags=["admin"], dependencies=[Depends(get_current_admin)])


def _apply_filters(
    query,
    q: str | None,
    pago: bool | None,
    zona: Zona | None,
    sexo: Sexo | None,
    tipo: TipoParticipante | None,
    asistio: bool | None = None,
    equipo_id: int | None = None,
):
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Camper.nombre.ilike(like),
                Camper.iglesia.ilike(like),
                Camper.folio.ilike(like),
                Camper.ciudad.ilike(like),
                Camper.telefono.ilike(like),
            )
        )
    if pago is not None:
        query = query.filter(Camper.pago_verificado == pago)
    if zona:
        query = query.filter(Camper.zona == zona)
    if sexo:
        query = query.filter(Camper.sexo == sexo)
    if tipo:
        query = query.filter(Camper.tipo == tipo)
    if asistio is not None:
        query = query.filter(Camper.asistio == asistio)
    if equipo_id is not None:
        # 0 es el valor centinela que usa el filtro del panel para "Sin equipo"
        # — no hay equipo con id 0 (autoincrement empieza en 1).
        query = query.filter(Camper.equipo_id.is_(None) if equipo_id == 0 else Camper.equipo_id == equipo_id)
    return query


@router.get("/registros", response_model=CamperListResponse)
def listar_registros(
    q: str | None = None,
    pago: bool | None = None,
    zona: Zona | None = None,
    sexo: Sexo | None = None,
    tipo: TipoParticipante | None = None,
    asistio: bool | None = None,
    equipo_id: int | None = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    db: Session = Depends(get_db),
):
    base = _apply_filters(db.query(Camper), q, pago, zona, sexo, tipo, asistio, equipo_id)
    total = base.count()
    items = (
        base.options(joinedload(Camper.equipo))
        .order_by(Camper.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return CamperListResponse(items=items, total=total, page=page, page_size=page_size)


@router.patch("/registros/{camper_id}/pago", response_model=CamperOut)
def actualizar_pago(
    camper_id: int,
    payload: PagoUpdate,
    admin: AdminUser = Depends(require_role(AdminRole.ADMIN, AdminRole.VERIFICADOR_PAGO)),
    db: Session = Depends(get_db),
):
    camper = db.get(Camper, camper_id)
    if not camper:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Registro no encontrado")

    camper.pago_verificado = payload.verificado
    camper.verificado_en = datetime.now(timezone.utc) if payload.verificado else None
    camper.verificado_por = admin.username if payload.verificado else None
    db.commit()
    db.refresh(camper)
    return camper


@router.patch("/registros/{camper_id}/asistencia", response_model=CamperOut)
def marcar_asistencia(
    camper_id: int,
    payload: AsistenciaUpdate,
    admin: AdminUser = Depends(require_role(AdminRole.ADMIN, AdminRole.RECEPCION)),
    db: Session = Depends(get_db),
):
    camper = db.get(Camper, camper_id)
    if not camper:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Registro no encontrado")

    camper.asistio = payload.asistio
    camper.asistio_en = datetime.now(timezone.utc) if payload.asistio else None
    camper.asistio_por = admin.username if payload.asistio else None
    db.commit()
    db.refresh(camper)
    return camper


@router.patch("/registros/{camper_id}/equipo", response_model=CamperOut)
def mover_de_equipo(
    camper_id: int,
    payload: EquipoAsignacion,
    _admin: AdminUser = Depends(require_role(AdminRole.ADMIN)),
    db: Session = Depends(get_db),
):
    """Mueve a alguien a mano desde /admin/equipos (o lo saca de su equipo con
    equipo_id=None). Marca equipo_fijado=True al asignar — es lo que hace que
    "Repartir a todos" respete este movimiento salvo que se pida lo contrario."""
    camper = db.get(Camper, camper_id)
    if not camper:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Registro no encontrado")

    if payload.equipo_id is not None and not db.get(Equipo, payload.equipo_id):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Equipo no encontrado")

    camper.equipo_id = payload.equipo_id
    camper.equipo_fijado = payload.equipo_id is not None
    db.commit()
    db.refresh(camper)
    return camper


@router.patch("/registros/{camper_id}", response_model=CamperOut)
def actualizar_registro(
    camper_id: int,
    payload: CamperUpdate,
    _admin: AdminUser = Depends(require_role(AdminRole.ADMIN)),
    db: Session = Depends(get_db),
):
    """Edita los datos de captura de un registro (nombre, edad, iglesia, zona,
    teléfono, bautismo, promoción, talla, fecha de pago). Deliberadamente no
    toca folio, comprobante, campos de auditoría de pago/asistencia ni
    equipo_id/equipo_fijado — eso se gestiona desde sus propios endpoints
    (/pago, /asistencia, /equipo, /admin/equipos). Tampoco vuelve a correr el
    reparto de equipos aunque cambien edad/sexo/iglesia/zona/bautismo: si
    quedó desbalanceado, "Repartir a todos" en /admin/equipos lo corrige."""
    camper = db.get(Camper, camper_id)
    if not camper:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Registro no encontrado")

    validar_datos_camper(
        tipo=payload.tipo,
        telefono=payload.telefono,
        bautizado=payload.bautizado,
        bautismo_mes=payload.bautismo_mes,
        bautismo_anio=payload.bautismo_anio,
        tiene_promocion=payload.tiene_promocion,
        promocion_detalle=payload.promocion_detalle,
        talla_camisa=payload.talla_camisa,
        talla_otra=payload.talla_otra,
    )

    datos = normalizar_datos_camper(
        nombre=payload.nombre,
        iglesia=payload.iglesia,
        edad=payload.edad,
        sexo=payload.sexo,
        zona=payload.zona,
        tipo=payload.tipo,
        telefono=payload.telefono,
        bautizado=payload.bautizado,
        bautismo_mes=payload.bautismo_mes,
        bautismo_anio=payload.bautismo_anio,
        fecha_pago=payload.fecha_pago,
        tiene_promocion=payload.tiene_promocion,
        promocion_detalle=payload.promocion_detalle,
        talla_camisa=payload.talla_camisa,
        talla_otra=payload.talla_otra,
    )
    for campo, valor in datos.items():
        setattr(camper, campo, valor)

    db.commit()
    db.refresh(camper)
    return camper


@router.delete("/registros/{camper_id}", status_code=status.HTTP_204_NO_CONTENT)
def borrar_registro(
    camper_id: int,
    _admin: AdminUser = Depends(require_role(AdminRole.ADMIN)),
    db: Session = Depends(get_db),
):
    camper = db.get(Camper, camper_id)
    if not camper:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Registro no encontrado")

    ticket_path = os.path.join(settings.tickets_dir, camper.ticket_path) if camper.ticket_path else None
    db.delete(camper)
    db.commit()

    if ticket_path:
        try:
            os.remove(ticket_path)
        except OSError:
            # El registro ya se borró; un comprobante huérfano en disco no es motivo
            # para fallar la petición (puede que el archivo ya no exista, por ejemplo).
            pass


@router.get("/registros/{camper_id}/ticket")
def ver_ticket(camper_id: int, db: Session = Depends(get_db)):
    camper = db.get(Camper, camper_id)
    if not camper:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Registro no encontrado")
    if not camper.ticket_path:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Este registro no tiene comprobante")

    path = os.path.join(settings.tickets_dir, camper.ticket_path)
    if not os.path.isfile(path):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Comprobante no encontrado")

    return FileResponse(path, media_type=camper.ticket_mime)


_MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

_EXPORT_HEADERS = [
    "Folio",
    "Nombre",
    "Tipo",
    "Teléfono",
    "Zona",
    "Iglesia",
    "Edad",
    "Sexo",
    "Talla",
    "Otra talla",
    "Fecha de pago",
    "Promoción",
    "Detalle promoción",
    "Bautizado",
    "Fecha de bautismo",
    "Comprobante",
    "Pago verificado",
    "Verificado por",
    "Asistió",
    "Asistió en",
    "Fecha registro",
    "Equipo",
]


def _sanitizar_celda(valor: str) -> str:
    # evita que Excel/LibreOffice interprete valores de registrantes como fórmulas (CWE-1236)
    if valor and valor[0] in ("=", "+", "-", "@"):
        return "'" + valor
    return valor


def _export_row(c: Camper) -> list:
    return [
        c.folio,
        _sanitizar_celda(c.nombre),
        "Consejero" if c.tipo == TipoParticipante.CONSEJERO else "Campero",
        _sanitizar_celda(c.telefono or ""),
        c.zona.value if c.zona else "",
        _sanitizar_celda(c.iglesia),
        c.edad,
        c.sexo.value,
        c.talla_camisa.value if c.talla_camisa else "",
        _sanitizar_celda(c.talla_otra or ""),
        c.fecha_pago.strftime("%Y-%m-%d") if c.fecha_pago else "",
        "Sí" if c.tiene_promocion else "No",
        _sanitizar_celda(c.promocion_detalle or ""),
        "Sí" if c.bautizado else "No",
        f"{_MESES[c.bautismo_mes - 1]} {c.bautismo_anio}" if c.bautismo_mes and c.bautismo_anio else "",
        "Sí" if c.ticket_path else "No",
        "Sí" if c.pago_verificado else "No",
        c.verificado_por or "",
        "Sí" if c.asistio else "No",
        c.asistio_en.strftime("%Y-%m-%d %H:%M") if c.asistio_en else "",
        c.created_at.strftime("%Y-%m-%d %H:%M"),
        _sanitizar_celda(c.equipo.nombre) if c.equipo else "Sin equipo",
    ]


def _export_rows(
    q: str | None,
    pago: bool | None,
    zona: Zona | None,
    sexo: Sexo | None,
    tipo: TipoParticipante | None,
    db: Session,
    asistio: bool | None = None,
    equipo_id: int | None = None,
) -> list[Camper]:
    base = _apply_filters(db.query(Camper), q, pago, zona, sexo, tipo, asistio, equipo_id)
    return base.options(joinedload(Camper.equipo)).order_by(Camper.created_at.desc()).all()


@router.get("/registros.csv")
def exportar_csv(
    q: str | None = None,
    pago: bool | None = None,
    zona: Zona | None = None,
    sexo: Sexo | None = None,
    tipo: TipoParticipante | None = None,
    asistio: bool | None = None,
    equipo_id: int | None = None,
    db: Session = Depends(get_db),
):
    rows = _export_rows(q, pago, zona, sexo, tipo, db, asistio, equipo_id)

    buffer = io.StringIO()
    buffer.write("﻿")  # BOM: para que Excel detecte UTF-8 y no manche los acentos
    writer = csv.writer(buffer)
    writer.writerow(_EXPORT_HEADERS)
    for c in rows:
        writer.writerow(_export_row(c))
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=registros_arraigados.csv"},
    )


@router.get("/registros.xlsx")
def exportar_xlsx(
    q: str | None = None,
    pago: bool | None = None,
    zona: Zona | None = None,
    sexo: Sexo | None = None,
    tipo: TipoParticipante | None = None,
    asistio: bool | None = None,
    equipo_id: int | None = None,
    db: Session = Depends(get_db),
):
    rows = _export_rows(q, pago, zona, sexo, tipo, db, asistio, equipo_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    header_font = Font(bold=True, color="FFF2D479")
    header_fill = PatternFill(fill_type="solid", fgColor="FF1B2B22")
    ws.append(_EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    widths = [len(h) for h in _EXPORT_HEADERS]
    for c in rows:
        row = _export_row(c)
        ws.append(row)
        for i, value in enumerate(row):
            widths[i] = max(widths[i], len(str(value)) if value is not None else 0)

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, width + 2), 45)

    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=registros_arraigados.xlsx"},
    )


@router.get("/stats/tallas", response_model=TallaStatsResponse)
def estadisticas_tallas(db: Session = Depends(get_db)):
    """Conteo de playeras por talla sobre TODOS los registros (no solo la
    página/filtro actual) — responde aunque el formulario público ya no pida
    talla, son datos ya capturados."""
    rows = (
        db.query(
            Camper.talla_camisa,
            func.count(Camper.id),
            func.sum(case((Camper.pago_verificado.is_(True), 1), else_=0)),
        )
        .group_by(Camper.talla_camisa)
        .all()
    )

    por_talla = {talla: (total, verificados or 0) for talla, total, verificados in rows if talla is not None}
    sin_talla = next((total for talla, total, _ in rows if talla is None), 0)

    items = [
        TallaStatsItem(talla=talla, total=por_talla[talla][0], verificados=por_talla[talla][1])
        for talla in TallaCamisa
        if talla in por_talla
    ]
    total_campers = sum(total for _, total, _ in rows)

    return TallaStatsResponse(items=items, sin_talla=sin_talla, total_campers=total_campers)


@router.get("/stats/comprobantes", response_model=ComprobanteStatsResponse)
def estadisticas_comprobantes(db: Session = Depends(get_db)):
    """Conteo de registros con/sin comprobante sobre TODOS los registros
    (no solo la página/filtro actual) — igual criterio que stats/tallas."""
    con_comprobante = db.query(func.count(Camper.id)).filter(Camper.ticket_path.isnot(None)).scalar() or 0
    sin_comprobante = db.query(func.count(Camper.id)).filter(Camper.ticket_path.is_(None)).scalar() or 0
    return ComprobanteStatsResponse(con_comprobante=con_comprobante, sin_comprobante=sin_comprobante)


@router.get("/stats/asistencia", response_model=AsistenciaStatsResponse)
def estadisticas_asistencia(db: Session = Depends(get_db)):
    """Conteo de check-in sobre TODOS los registros — lo consume la pantalla
    de recepción para sus tres StatTile de arriba."""
    total = db.query(func.count(Camper.id)).scalar() or 0
    asistieron = db.query(func.count(Camper.id)).filter(Camper.asistio.is_(True)).scalar() or 0
    return AsistenciaStatsResponse(total=total, asistieron=asistieron, faltan=total - asistieron)


@router.patch("/settings", response_model=AppSettingsOut)
def actualizar_settings(
    payload: AppSettingsUpdate,
    _admin: AdminUser = Depends(require_role(AdminRole.ADMIN)),
    db: Session = Depends(get_db),
):
    updates = payload.model_dump(exclude_unset=True)
    row = db.get(AppSettings, 1)
    if not row:
        # No debería pasar (seed_settings() la crea al arrancar), pero si la fila
        # sembrada faltara por alguna razón, se crea aquí en vez de fallar.
        row = AppSettings(id=1)
        db.add(row)
    for field, value in updates.items():
        setattr(row, field, value)
    db.commit()
    db.refresh(row)
    return AppSettingsOut(
        show_shirt_size=row.show_shirt_size,
        precio_mxn=row.precio_mxn,
        pedir_comprobante=row.pedir_comprobante,
        registro_abierto=row.registro_abierto,
    )


# ---- Gestión de cuentas de admin (todas exigen rol Admin) ----

usuarios_router = APIRouter(
    prefix="/api/admin/usuarios",
    tags=["admin-usuarios"],
    dependencies=[Depends(require_role(AdminRole.ADMIN))],
)


@usuarios_router.get("", response_model=list[AdminUserOut])
def listar_usuarios(db: Session = Depends(get_db)):
    return db.query(AdminUser).order_by(AdminUser.created_at.asc()).all()


@usuarios_router.post("", response_model=AdminUserOut, status_code=status.HTTP_201_CREATED)
def crear_usuario(payload: AdminUserCreate, db: Session = Depends(get_db)):
    if db.query(AdminUser).filter(AdminUser.username == payload.username).first():
        raise HTTPException(status.HTTP_409_CONFLICT, "Ese nombre de usuario ya existe.")

    admin = AdminUser(
        username=payload.username,
        password_hash=hash_password(payload.password),
        role=payload.role,
    )
    db.add(admin)
    db.commit()
    db.refresh(admin)
    return admin


@usuarios_router.patch("/{admin_id}", response_model=AdminUserOut)
def actualizar_usuario(
    admin_id: int,
    payload: AdminUserUpdate,
    admin: AdminUser = Depends(require_role(AdminRole.ADMIN)),
    db: Session = Depends(get_db),
):
    target = db.get(AdminUser, admin_id)
    if not target:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Cuenta no encontrada")

    if payload.role is not None:
        if target.id == admin.id:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "No puedes cambiar tu propio rol.")
        target.role = payload.role

    if payload.password:
        target.password_hash = hash_password(payload.password)

    db.commit()
    db.refresh(target)
    return target


@usuarios_router.delete("/{admin_id}", status_code=status.HTTP_204_NO_CONTENT)
def borrar_usuario(
    admin_id: int,
    admin: AdminUser = Depends(require_role(AdminRole.ADMIN)),
    db: Session = Depends(get_db),
):
    target = db.get(AdminUser, admin_id)
    if not target:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Cuenta no encontrada")

    if target.id == admin.id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No puedes borrar tu propia cuenta.")

    if target.role == AdminRole.ADMIN:
        admins_restantes = db.query(AdminUser).filter(AdminUser.role == AdminRole.ADMIN).count()
        if admins_restantes <= 1:
            raise HTTPException(status.HTTP_409_CONFLICT, "Debe existir al menos un administrador.")

    db.delete(target)
    db.commit()
